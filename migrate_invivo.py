"""
Reload the `in_vivo` table from a curated Excel workbook.

Workflow:
  1. Inspect live `in_vivo` schema (column list + types).
  2. Build header mapping from spreadsheet → DB columns (case-insensitive after
     stripping whitespace and replacing space/period/hyphen with underscore;
     spreadsheet column index 3 has a blank header → renamed to "Type").
  3. Create `in_vivo_new` via `LIKE in_vivo INCLUDING ALL` (clones types,
     defaults, indexes).
  4. Stream the spreadsheet, transform Type values, coerce Excel error cells
     (#DIV/0! etc.) to NULL on numeric columns, and COPY into the staging
     table in a single transaction.
  5. Run sanity queries; abort if any fail.
  6. With --swap, perform the atomic rename:
        BEGIN;
        ALTER TABLE in_vivo RENAME TO in_vivo_old_<ts>;
        ALTER TABLE in_vivo_new RENAME TO in_vivo;
        COMMIT;
     Indexes follow the table on rename in PostgreSQL >= 9.4.

Re-runnable: drops and recreates `in_vivo_new` on each invocation.

Usage:
    python migrate_invivo.py                                  # stage + verify (no swap)
    python migrate_invivo.py --swap                           # also perform the rename swap
    python migrate_invivo.py --xlsx Database_invivo_v2.xlsx   # use a different file

For an in_vitro variant, fork this file and:
    - point TABLE at the in_vitro table name
    - update TYPE_VALUE_MAP to whatever the in_vitro spreadsheet uses
    - revisit the col-3 blank-header override if the in_vitro sheet doesn't
      have a similar quirk
"""
from __future__ import annotations

import argparse
import csv
import io
import os
import sys
import time
from datetime import datetime, timezone

import openpyxl
import psycopg2
from dotenv import load_dotenv

TABLE = "in_vivo"
DEFAULT_XLSX = "Database_invivo_corrected.xlsx"
DEFAULT_SHEET = "Sheet1"

TYPE_VALUE_MAP = {
    "By HILIC_Negative": "by_negative",
    "By RPLC_Positive":  "by_positive",
    "By Name":           "by_name",
}

# Spreadsheet column index → header override. The shipped in_vivo workbook has
# a blank header at index 3 whose values are the Type label.
HEADER_OVERRIDES = {3: "Type"}


def normalize_header(name: str | None) -> str | None:
    if name is None:
        return None
    s = str(name).strip()
    return s.replace(" ", "_").replace(".", "_").replace("-", "_")


def db_columns_in_order(cur, table: str) -> tuple[list[str], list[str]]:
    cur.execute(
        """SELECT column_name, data_type FROM information_schema.columns
           WHERE table_name=%s ORDER BY ordinal_position""",
        (table,),
    )
    rows = cur.fetchall()
    return [r[0] for r in rows], [r[1] for r in rows]


def build_ss_to_db_index_map(ss_headers, db_cols):
    """Length=len(db_cols); each entry = spreadsheet col index that feeds it."""
    db_lc_to_idx = {c.lower(): i for i, c in enumerate(db_cols)}
    ss_to_db: list[int | None] = [None] * len(db_cols)
    for ss_idx, h in enumerate(ss_headers):
        norm = normalize_header(HEADER_OVERRIDES.get(ss_idx, h))
        if norm is None:
            continue
        db_idx = db_lc_to_idx.get(norm.lower())
        if db_idx is None:
            raise RuntimeError(f"Spreadsheet col {ss_idx} {h!r} -> {norm!r} has no matching DB column")
        ss_to_db[db_idx] = ss_idx
    missing = [db_cols[i] for i, v in enumerate(ss_to_db) if v is None]
    if missing:
        raise RuntimeError(f"DB columns not provided by spreadsheet: {missing}")
    return ss_to_db


def stage_rows(conn, xlsx: str, sheet: str, db_cols: list[str],
               db_types: list[str], ss_to_db: list[int]) -> int:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[sheet]

    quoted_cols = ", ".join(f'"{c}"' for c in db_cols)
    type_db_idx = db_cols.index("Type")
    name_db_idx = db_cols.index("name")

    # Numeric DB columns need string-coercion for Excel errors like #DIV/0!.
    NUMERIC_TYPES = {"double precision", "numeric", "real", "integer", "bigint", "smallint"}
    numeric_idxs = {i for i, t in enumerate(db_types) if t in NUMERIC_TYPES}

    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)

    rows_in = 0
    rows_out = 0
    skipped_null = 0
    coerced_to_null = 0
    unmapped: list[tuple[int, object]] = []
    for ss_row in ws.iter_rows(min_row=2, values_only=True):
        rows_in += 1
        out_row: list = [None] * len(db_cols)
        for db_idx, ss_idx in enumerate(ss_to_db):
            out_row[db_idx] = ss_row[ss_idx]

        if out_row[name_db_idx] is None:
            skipped_null += 1
            continue

        raw_type = out_row[type_db_idx]
        mapped = TYPE_VALUE_MAP.get(raw_type) if raw_type is not None else None
        if mapped is None:
            unmapped.append((rows_in, raw_type))
            continue
        out_row[type_db_idx] = mapped

        for ni in numeric_idxs:
            v = out_row[ni]
            if v is None or not isinstance(v, str):
                continue
            try:
                out_row[ni] = float(v)
            except ValueError:
                out_row[ni] = None
                coerced_to_null += 1

        writer.writerow(["" if v is None else v for v in out_row])
        rows_out += 1

    if unmapped:
        print(f"ERROR: {len(unmapped)} rows had unmapped Type values; first 5: {unmapped[:5]}", file=sys.stderr)
        raise RuntimeError("Unmapped Type values — aborting before any DB write.")

    print(f"  read {rows_in} rows; skipped {skipped_null} (null name); staged {rows_out}; coerced {coerced_to_null} non-numeric cells -> NULL")

    buf.seek(0)
    with conn.cursor() as cur:
        cur.copy_expert(
            f"COPY {TABLE}_new ({quoted_cols}) FROM STDIN WITH (FORMAT csv, NULL '')",
            buf,
        )
    return rows_out


def verify_staging(conn) -> None:
    """Log post-load shape so a human can sanity-check before committing to swap.
    No hard assertions — counts shift between refreshes by design."""
    with conn.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM {TABLE}_new;")
        print(f"  row_count = {cur.fetchone()[0]}")

        cur.execute(f'SELECT "Type", COUNT(*) FROM {TABLE}_new GROUP BY "Type" ORDER BY 2 DESC;')
        print(f"  type_distribution = {dict(cur.fetchall())}")

        cur.execute(f'SELECT COUNT(*) FROM {TABLE}_new WHERE "name" IS NULL;')
        null_names = cur.fetchone()[0]
        if null_names != 0:
            raise RuntimeError(f"Found {null_names} null-name rows; expected 0")
        print(f"  null_names = 0 OK")

        cur.execute(f'SELECT "name","mz","rt","Type" FROM {TABLE}_new ORDER BY "name" LIMIT 3;')
        print(f"  sample first 3 by name:")
        for r in cur.fetchall():
            print(f"    {r}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    ap.add_argument("--xlsx", default=DEFAULT_XLSX, help=f"Path to xlsx (default: {DEFAULT_XLSX})")
    ap.add_argument("--sheet", default=DEFAULT_SHEET, help=f"Sheet name (default: {DEFAULT_SHEET})")
    ap.add_argument("--swap", action="store_true", help="Perform the atomic rename swap after staging.")
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: spreadsheet not found at {args.xlsx}", file=sys.stderr); return 2

    load_dotenv()
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("ERROR: DATABASE_URL not set", file=sys.stderr); return 2

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    print(f"=== migrate_{TABLE} run @ {ts} (xlsx={args.xlsx}) ===")

    conn = psycopg2.connect(db_url)
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            db_cols, db_types = db_columns_in_order(cur, TABLE)
            print(f"  DB has {len(db_cols)} columns in {TABLE}")

        wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
        ws = wb[args.sheet]
        ss_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        wb.close()
        ss_to_db = build_ss_to_db_index_map(ss_headers, db_cols)
        print(f"  built ss->db map: {len(ss_to_db)} columns")

        with conn.cursor() as cur:
            print(f"  dropping & recreating {TABLE}_new (LIKE {TABLE} INCLUDING ALL)")
            cur.execute(f"DROP TABLE IF EXISTS {TABLE}_new;")
            cur.execute(f"CREATE TABLE {TABLE}_new (LIKE {TABLE} INCLUDING ALL);")

        t0 = time.time()
        print(f"  streaming xlsx -> COPY {TABLE}_new")
        stage_rows(conn, args.xlsx, args.sheet, db_cols, db_types, ss_to_db)
        print(f"  staged in {time.time()-t0:.1f}s")

        print("  verifying staging:")
        verify_staging(conn)

        conn.commit()
        print(f"  committed {TABLE}_new")

        if args.swap:
            with conn.cursor() as cur:
                old_name = f"{TABLE}_old_{ts}"
                print(f"  performing atomic rename swap -> {old_name}")
                cur.execute(f'ALTER TABLE {TABLE} RENAME TO "{old_name}";')
                cur.execute(f'ALTER TABLE {TABLE}_new RENAME TO {TABLE};')
            conn.commit()
            print(f"  SWAP COMPLETE. Rollback table: {old_name}")
        else:
            print("  --swap not specified; staging only. Re-run with --swap to perform the rename.")
        return 0
    except Exception as e:
        conn.rollback()
        print(f"ERROR: {e}", file=sys.stderr)
        return 1
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
